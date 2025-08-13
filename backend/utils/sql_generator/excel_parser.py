"""
Excel file parser for extracting mapping information from multiple sheets
"""
import openpyxl
from typing import Dict, List, Any, Optional
import json


class ExcelMappingParser:
    """
    Parse Excel files with multiple sheets containing vague mapping information
    """
    
    def __init__(self):
        self.parsed_sheets = {}
    
    async def parse_excel_file(self, file_path: str) -> Dict[str, Any]:
        """
        Parse Excel file and extract all sheet data for AI processing
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary with sheet names as keys and extracted data as values
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            for sheet_name in workbook.sheetnames:
                sheet_data = await self._extract_sheet_data(workbook[sheet_name])
                self.parsed_sheets[sheet_name] = sheet_data
            
            return self.parsed_sheets
        
        except Exception as e:
            raise Exception(f"Failed to parse Excel file: {str(e)}")
    
    async def parse_excel_bytes(self, file_bytes: bytes, filename: str) -> Dict[str, Any]:
        """
        Parse Excel file from bytes (for uploaded files)
        
        Args:
            file_bytes: Excel file content as bytes
            filename: Original filename for context
            
        Returns:
            Dictionary with sheet names as keys and extracted data as values
        """
        try:
            from io import BytesIO
            
            # Load workbook from bytes
            workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
            
            for sheet_name in workbook.sheetnames:
                sheet_data = await self._extract_sheet_data(workbook[sheet_name])
                self.parsed_sheets[sheet_name] = sheet_data
            
            # Add metadata
            self.parsed_sheets["_metadata"] = {
                "filename": filename,
                "total_sheets": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames
            }
            
            return self.parsed_sheets
        
        except Exception as e:
            raise Exception(f"Failed to parse Excel bytes: {str(e)}")
    
    async def _extract_sheet_data(self, worksheet) -> Dict[str, Any]:
        """
        Extract all meaningful data from a single worksheet
        """
        sheet_data = {
            "name": worksheet.title,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "cells_with_data": [],
            "tables": [],
            "headers": [],
            "raw_text": []
        }
        
        # Extract all non-empty cells
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_info = {
                        "row": cell.row,
                        "column": cell.column,
                        "column_letter": cell.column_letter,
                        "value": str(cell.value),
                        "data_type": str(type(cell.value).__name__)
                    }
                    sheet_data["cells_with_data"].append(cell_info)
                    sheet_data["raw_text"].append(str(cell.value))
        
        # Try to identify table structures
        sheet_data["tables"] = await self._identify_tables(worksheet)
        
        # Try to identify headers/sections
        sheet_data["headers"] = await self._identify_headers(worksheet)
        
        # Create a text representation for AI
        sheet_data["text_representation"] = await self._create_text_representation(sheet_data)
        
        return sheet_data
    
    async def _identify_tables(self, worksheet) -> List[Dict]:
        """
        Attempt to identify table-like structures in the sheet
        """
        tables = []
        
        # Look for consecutive rows with data that might represent tables
        for row_num in range(1, worksheet.max_row + 1):
            row_data = []
            for col_num in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    row_data.append(str(cell_value))
                else:
                    row_data.append("")
            
            # If row has multiple non-empty cells, it might be a table row
            non_empty_count = len([cell for cell in row_data if cell.strip()])
            if non_empty_count >= 2:
                tables.append({
                    "row": row_num,
                    "data": row_data,
                    "non_empty_cells": non_empty_count
                })
        
        return tables
    
    async def _identify_headers(self, worksheet) -> List[Dict]:
        """
        Identify potential headers or section titles
        """
        headers = []
        
        for row_num in range(1, worksheet.max_row + 1):
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    cell_value = str(cell.value)
                    
                    # Heuristics for identifying headers
                    is_potential_header = (
                        len(cell_value) > 3 and  # Not too short
                        (cell.font and cell.font.bold) or  # Bold formatting
                        cell_value.isupper() or  # All caps
                        ":" in cell_value or  # Contains colon
                        any(keyword in cell_value.lower() for keyword in 
                            ["table", "field", "column", "mapping", "source", "target", "join"])
                    )
                    
                    if is_potential_header:
                        headers.append({
                            "row": row_num,
                            "column": col_num,
                            "value": cell_value,
                            "formatted": bool(cell.font and cell.font.bold)
                        })
        
        return headers
    
    async def _create_text_representation(self, sheet_data: Dict) -> str:
        """
        Create a text representation of the sheet for AI processing
        """
        text_parts = [
            f"SHEET: {sheet_data['name']}",
            f"DIMENSIONS: {sheet_data['max_row']} rows x {sheet_data['max_column']} columns",
            ""
        ]
        
        # Add headers section
        if sheet_data["headers"]:
            text_parts.append("IDENTIFIED HEADERS/SECTIONS:")
            for header in sheet_data["headers"]:
                text_parts.append(f"  Row {header['row']}: {header['value']}")
            text_parts.append("")
        
        # Add table-like structures
        if sheet_data["tables"]:
            text_parts.append("POTENTIAL TABLE DATA:")
            for i, table_row in enumerate(sheet_data["tables"][:10]):  # Limit to first 10 rows
                non_empty_data = [cell for cell in table_row["data"] if cell.strip()]
                text_parts.append(f"  Row {table_row['row']}: {' | '.join(non_empty_data)}")
            text_parts.append("")
        
        # Add all raw text for context
        text_parts.append("ALL TEXT CONTENT:")
        unique_texts = list(set(sheet_data["raw_text"]))  # Remove duplicates
        text_parts.extend([f"  {text}" for text in unique_texts[:50]])  # Limit to first 50 unique texts
        
        return "\n".join(text_parts)
    
    def get_consolidated_text_for_ai(self) -> str:
        """
        Create a consolidated text representation of all sheets for AI processing
        """
        if not self.parsed_sheets:
            return "No sheet data available"
        
        consolidated_parts = [
            "EXCEL MAPPING DOCUMENT ANALYSIS",
            "=" * 50,
            ""
        ]
        
        # Add metadata if available
        if "_metadata" in self.parsed_sheets:
            metadata = self.parsed_sheets["_metadata"]
            consolidated_parts.extend([
                f"FILENAME: {metadata.get('filename', 'Unknown')}",
                f"TOTAL SHEETS: {metadata.get('total_sheets', 0)}",
                f"SHEET NAMES: {', '.join(metadata.get('sheet_names', []))}",
                ""
            ])
        
        # Add each sheet's text representation
        for sheet_name, sheet_data in self.parsed_sheets.items():
            if sheet_name != "_metadata":
                consolidated_parts.extend([
                    f"{'='*20} SHEET: {sheet_name} {'='*20}",
                    sheet_data.get("text_representation", "No data"),
                    "",
                    "-" * 50,
                    ""
                ])
        
        return "\n".join(consolidated_parts)
